from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl.utils import get_column_letter

from src.utils.sheet_utils import safe_sheet_name


logger = logging.getLogger(__name__)


def save_to_pickle(sheets_dict: Dict[str, pd.DataFrame], output_path: Path) -> None:
    logger.info("Saving pickle: %s", output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    pd.to_pickle(sheets_dict, output_path)


def _autofit_openpyxl_sheet(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    """Auto-fit columns using openpyxl like in the notebook."""
    try:
        ws = writer.book[sheet_name]
        df_reset = df.reset_index() if df.index.names != [None] else df
        for idx, column in enumerate(df_reset.columns, start=1):
            col_letter = get_column_letter(idx)
            values = df_reset[column].astype(str).tolist()
            max_len = max([len(str(column))] + [len(v) for v in values]) if values else len(str(column))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
    except Exception:
        logger.debug("Autofit failed for sheet %s", sheet_name, exc_info=True)


def save_to_excel_multisheet(sheets_dict: Dict[str, pd.DataFrame], output_path: Path) -> None:
    logger.info("Saving Excel (multi-sheet): %s", output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for raw_name, df in sheets_dict.items():
            name = safe_sheet_name(raw_name)
            df.to_excel(writer, sheet_name=name)
            _autofit_openpyxl_sheet(writer, name, df)


def save_to_excel_single_report(
    sheets_dict: Dict[str, pd.DataFrame], output_path: Path, sheet_name: str = "Global"
) -> None:
    """Write all DataFrames into one sheet using openpyxl engine.
    
    Uses simple title rows and the same column auto-fit logic as the notebook.
    """
    logger.info("Saving Excel (single-sheet): %s", output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        ws_name = safe_sheet_name(sheet_name)
        # Create sheet and register
        ws = writer.book.create_sheet(ws_name)
        writer.sheets[ws_name] = ws
        row_cursor = 1  # openpyxl is 1-based for cell addressing, pandas uses 0-based startrow
        for title, df in sheets_dict.items():
            # Write title (plain text to match minimal notebook style)
            ws.cell(row=row_cursor, column=1, value=str(title))
            row_cursor += 1
            # Write the DataFrame starting at row_cursor-1 for pandas startrow
            df_to_write = df.copy()
            df_to_write.to_excel(writer, sheet_name=ws_name, startrow=row_cursor - 1, index=True)
            # Auto-fit columns for current written block
            _autofit_openpyxl_sheet(writer, ws_name, df_to_write)
            row_cursor += len(df_to_write.index) + 4  # space between tables


